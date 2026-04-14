#!/usr/bin/env python3
"""Parse TestIT xlsx export into normalized tests.json.

Layout:
  A = id, C = title, F = step action, H = expected result.
  Row with non-empty A and C opens a new test; subsequent rows with empty A/C
  are steps of the current test.
"""
import argparse
import json
import re
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value)
    if "<" in text and ">" in text:
        text = BeautifulSoup(text, "html.parser").get_text(" ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    tests: list[dict] = []
    current: dict | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        # pad row to at least 8 cols (A..H)
        row = list(row) + [None] * (8 - len(row))
        a, c, f, h = row[0], row[2], row[5], row[7]
        id_val = clean(a)
        title = clean(c)
        action = clean(f)
        expected = clean(h)

        if id_val and title:
            if current:
                tests.append(current)
            current = {"id": id_val, "title": title, "steps": []}
            if action or expected:
                current["steps"].append({"action": action, "expected": expected})
        elif current and (action or expected):
            current["steps"].append({"action": action, "expected": expected})

    if current:
        tests.append(current)

    return tests


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--out", type=Path, default=Path("tests.json"))
    args = ap.parse_args()

    tests = parse(args.xlsx)
    args.out.write_text(json.dumps(tests, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"extracted {len(tests)} tests → {args.out}")


if __name__ == "__main__":
    main()
